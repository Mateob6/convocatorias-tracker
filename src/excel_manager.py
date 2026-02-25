"""Gestion del archivo Excel de convocatorias."""

import logging
from collections import Counter
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src.models import COLUMNAS, Oportunidad

logger = logging.getLogger(__name__)

# Style constants
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

VENCIDA_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
ALTA_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
URGENTE_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

COLUMN_WIDTHS = {
    1: 5,    # ID
    2: 14,   # Fecha_deteccion
    3: 48,   # Nombre
    4: 30,   # Entidad
    5: 16,   # Tipo
    6: 18,   # Fuente
    7: 45,   # URL
    8: 14,   # Fecha_apertura
    9: 14,   # Fecha_cierre
    10: 22,  # Monto
    11: 50,  # Requisitos_clave
    12: 40,  # Documentos_necesarios
    13: 12,  # Relevancia
    14: 16,  # Estado
    15: 50,  # Notas
}


class ExcelManager:
    """Manages the convocatorias Excel workbook."""

    def __init__(self, filepath: str):
        self.filepath = Path(filepath)
        self.wb: Optional[Workbook] = None
        self._existing_keys: Optional[set] = None

    def open_or_create(self) -> None:
        """Open existing workbook or create new one with headers and formatting."""
        if self.filepath.exists():
            self.wb = load_workbook(str(self.filepath))
            logger.info(f"Opened existing Excel: {self.filepath}")
        else:
            self.filepath.parent.mkdir(parents=True, exist_ok=True)
            self.wb = Workbook()
            self._setup_main_sheet()
            self._setup_summary_sheet()
            logger.info(f"Created new Excel: {self.filepath}")

    def _setup_main_sheet(self) -> None:
        """Create 'Convocatorias' sheet with formatted headers."""
        ws = self.wb.active
        ws.title = "Convocatorias"

        for col_idx, header in enumerate(COLUMNAS, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGN
            cell.border = THIN_BORDER

        for col_idx, width in COLUMN_WIDTHS.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNAS))}1"
        ws.freeze_panes = "A2"

    def _setup_summary_sheet(self) -> None:
        """Create 'Resumen' sheet."""
        if "Resumen" in self.wb.sheetnames:
            return
        ws = self.wb.create_sheet("Resumen")
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["D"].width = 25
        ws.column_dimensions["E"].width = 12
        ws.column_dimensions["G"].width = 25
        ws.column_dimensions["H"].width = 12

    def get_next_id(self) -> int:
        """Return max existing ID + 1, or 1 if no rows exist."""
        ws = self.wb["Convocatorias"]
        max_id = 0
        for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
            if row[0] is not None:
                try:
                    max_id = max(max_id, int(row[0]))
                except (ValueError, TypeError):
                    pass
        return max_id + 1

    def _load_existing_keys(self) -> set:
        """Load deduplication keys from existing rows."""
        if self._existing_keys is not None:
            return self._existing_keys

        self._existing_keys = set()
        ws = self.wb["Convocatorias"]
        for row in ws.iter_rows(min_row=2, max_col=len(COLUMNAS), values_only=True):
            if row[2] is not None:  # nombre column
                op = Oportunidad.from_row(list(row))
                self._existing_keys.add(op.dedup_key())
        return self._existing_keys

    def add_oportunidad(self, op: Oportunidad) -> bool:
        """
        Append a new Oportunidad as a row. Returns True if added, False if duplicate.
        """
        existing = self._load_existing_keys()
        key = op.dedup_key()
        if key in existing:
            logger.debug(f"Duplicate skipped: {op.nombre}")
            return False

        if op.id is None:
            op.id = self.get_next_id()

        ws = self.wb["Convocatorias"]
        row_data = op.to_row()
        ws.append(row_data)

        row_idx = ws.max_row
        for col_idx in range(1, len(COLUMNAS) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)

        existing.add(key)
        return True

    def add_oportunidades(self, ops: List[Oportunidad]) -> int:
        """Batch add with deduplication. Returns count of actually added."""
        added = 0
        for op in ops:
            if self.add_oportunidad(op):
                added += 1
        return added

    def get_all_oportunidades(self) -> List[Oportunidad]:
        """Read all opportunities from the main sheet."""
        ws = self.wb["Convocatorias"]
        results = []
        for row in ws.iter_rows(min_row=2, max_col=len(COLUMNAS), values_only=True):
            if row[0] is not None:
                results.append(Oportunidad.from_row(list(row)))
        return results

    def update_estados(self) -> int:
        """Mark expired entries as 'Vencida'. Returns count updated."""
        ws = self.wb["Convocatorias"]
        updated = 0
        today = date.today()
        estado_col = COLUMNAS.index("Estado") + 1  # 14
        cierre_col = COLUMNAS.index("Fecha_cierre") + 1  # 9

        for row_idx in range(2, ws.max_row + 1):
            cierre_val = ws.cell(row=row_idx, column=cierre_col).value
            estado_val = ws.cell(row=row_idx, column=estado_col).value

            if cierre_val is None:
                continue

            if isinstance(cierre_val, datetime):
                cierre_date = cierre_val.date()
            elif isinstance(cierre_val, date):
                cierre_date = cierre_val
            else:
                continue

            if cierre_date < today and estado_val not in ("Vencida", "Aplicada", "En proceso"):
                ws.cell(row=row_idx, column=estado_col, value="Vencida")
                updated += 1

        if updated:
            logger.info(f"Marked {updated} opportunities as Vencida")
        return updated

    def sort_by_fecha_cierre(self) -> None:
        """Sort data rows by Fecha_cierre ascending. Rows with no date go to the end."""
        ws = self.wb["Convocatorias"]
        if ws.max_row < 3:
            return

        # Read all data rows, ensuring we get exactly len(COLUMNAS) values per row
        data_rows = []
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(COLUMNAS)):
            row_values = [cell.value for cell in row]
            # Pad with None if row is shorter than expected
            while len(row_values) < len(COLUMNAS):
                row_values.append(None)
            data_rows.append(row_values)

        far_future = date(9999, 12, 31)

        def sort_key(row_vals):
            cierre = row_vals[8]  # Fecha_cierre index
            if cierre is None:
                return far_future
            if isinstance(cierre, datetime):
                return cierre.date()
            if isinstance(cierre, date):
                return cierre
            return far_future

        data_rows.sort(key=sort_key)

        # Clear all data cells first to prevent stale values
        for r_idx in range(2, ws.max_row + 1):
            for c_idx in range(1, len(COLUMNAS) + 1):
                ws.cell(row=r_idx, column=c_idx).value = None

        # Write sorted data back
        for r_idx, row_vals in enumerate(data_rows, 2):
            for c_idx, val in enumerate(row_vals, 1):
                ws.cell(row=r_idx, column=c_idx).value = val

    def apply_conditional_formatting(self) -> None:
        """Apply color coding based on estado, relevancia, and fecha_cierre."""
        ws = self.wb["Convocatorias"]
        today = date.today()
        urgente_limit = today + timedelta(days=15)

        estado_col = COLUMNAS.index("Estado") + 1
        relevancia_col = COLUMNAS.index("Relevancia") + 1
        cierre_col = COLUMNAS.index("Fecha_cierre") + 1

        for row_idx in range(2, ws.max_row + 1):
            estado = ws.cell(row=row_idx, column=estado_col).value
            relevancia = ws.cell(row=row_idx, column=relevancia_col).value
            cierre_val = ws.cell(row=row_idx, column=cierre_col).value

            if estado == "Vencida":
                for col_idx in range(1, len(COLUMNAS) + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = VENCIDA_FILL
                continue

            if relevancia == "Alta":
                for col_idx in range(1, len(COLUMNAS) + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = ALTA_FILL

            if cierre_val is not None:
                cierre_date = cierre_val
                if isinstance(cierre_date, datetime):
                    cierre_date = cierre_date.date()
                if isinstance(cierre_date, date) and today <= cierre_date <= urgente_limit:
                    ws.cell(row=row_idx, column=cierre_col).fill = URGENTE_FILL

    def update_summary_sheet(self) -> None:
        """Rebuild the Resumen sheet with current counts."""
        if "Resumen" not in self.wb.sheetnames:
            self._setup_summary_sheet()

        ws = self.wb["Resumen"]
        # Clear existing content
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=10):
            for cell in row:
                cell.value = None

        ops = self.get_all_oportunidades()

        # Title
        title_cell = ws.cell(row=1, column=1, value="Resumen de Convocatorias")
        title_cell.font = Font(bold=True, size=14, color="2F5496")

        ws.cell(row=2, column=1, value=f"Ultima actualizacion: {date.today().strftime('%d/%m/%Y')}")
        ws.cell(row=2, column=1).font = Font(italic=True, color="666666")

        ws.cell(row=3, column=1, value=f"Total de convocatorias: {len(ops)}")
        ws.cell(row=3, column=1).font = Font(bold=True)

        # Count by Tipo
        row = 5
        ws.cell(row=row, column=1, value="Por Tipo")
        ws.cell(row=row, column=1).font = Font(bold=True, size=12)
        ws.cell(row=row, column=2, value="Cantidad")
        ws.cell(row=row, column=2).font = Font(bold=True, size=12)
        tipo_counts = Counter(op.tipo for op in ops)
        for tipo, count in sorted(tipo_counts.items()):
            row += 1
            ws.cell(row=row, column=1, value=tipo)
            ws.cell(row=row, column=2, value=count)

        # Count by Estado
        row += 2
        ws.cell(row=row, column=1, value="Por Estado")
        ws.cell(row=row, column=1).font = Font(bold=True, size=12)
        ws.cell(row=row, column=2, value="Cantidad")
        ws.cell(row=row, column=2).font = Font(bold=True, size=12)
        estado_counts = Counter(op.estado for op in ops)
        for estado, count in sorted(estado_counts.items()):
            row += 1
            ws.cell(row=row, column=1, value=estado)
            ws.cell(row=row, column=2, value=count)

        # Count by Relevancia
        row += 2
        ws.cell(row=row, column=1, value="Por Relevancia")
        ws.cell(row=row, column=1).font = Font(bold=True, size=12)
        ws.cell(row=row, column=2, value="Cantidad")
        ws.cell(row=row, column=2).font = Font(bold=True, size=12)
        rel_counts = Counter(op.relevancia for op in ops)
        for rel, count in sorted(rel_counts.items()):
            row += 1
            ws.cell(row=row, column=1, value=rel)
            ws.cell(row=row, column=2, value=count)

        # Upcoming deadlines
        row += 2
        ws.cell(row=row, column=1, value="Proximas a vencer (15 dias)")
        ws.cell(row=row, column=1).font = Font(bold=True, size=12)
        today = date.today()
        limit = today + timedelta(days=15)
        upcoming = [
            op for op in ops
            if op.fecha_cierre and today <= op.fecha_cierre <= limit
            and op.estado not in ("Vencida", "Aplicada")
        ]
        if upcoming:
            for op in sorted(upcoming, key=lambda o: o.fecha_cierre):
                row += 1
                ws.cell(row=row, column=1, value=op.nombre)
                ws.cell(row=row, column=2, value=op.fecha_cierre)
                ws.cell(row=row, column=2).fill = URGENTE_FILL
        else:
            row += 1
            ws.cell(row=row, column=1, value="Ninguna en los proximos 15 dias")

    def save(self) -> None:
        """Save workbook to filepath."""
        self.wb.save(str(self.filepath))
        logger.info(f"Excel saved to {self.filepath}")
